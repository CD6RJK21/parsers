from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import openpyxl
import time
from multiprocessing import Pool


def parse(i):
    m = i + 10

    with open('id3.txt', 'r') as file:
        lines = [line.replace('\n', '') for line in file.readlines()]

    filename = 'declaracii_multi.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['decla']



    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument("user-data-dir=selenium")
    # driver = webdriver.Chrome(chrome_options=chrome_options)
    driver = webdriver.Chrome()
    m = i + m
    while i <= m:
        line = lines[i]
        url = "https://pub.fsa.gov.ru/rds/declaration/view/{}/applicant".format(
            line)
        driver.get(url)
        try:
            text = WebDriverWait(driver, 2).until(
                EC.visibility_of_element_located(
                    (By.XPATH,
                     "/html/body/fgis-root/div/fgis-rds-view-declaration/fgis-rds-view-declaration-toolbar/div"))).text
        except TimeoutException as te:
            time.sleep(5)
            continue
        try:
            text = driver.find_element_by_xpath(
                '/html/body/fgis-root/div/fgis-rds-view-declaration/div/div/div/div/fgis-rds-view-application-applicant/fgis-card-block-wrapper/fgis-card-block/div/div[2]/div/fgis-card-info-row/div[2]/span').text
            ws['A{}'.format(str(i + 2))] = text
        except BaseException as be:
            pass
        try:
            text = driver.find_element_by_xpath(
                '/html/body/fgis-root/div/fgis-rds-view-declaration/div/div/div/div/fgis-rds-view-application-applicant/fgis-card-block-wrapper/fgis-card-block/div/div[2]/div/fgis-rds-view-contacts/fgis-card-row-spoiler/div[2]/fgis-card-edit-row-two-columns/fgis-card-info-row[1]/div[2]/p').text
            ws['B{}'.format(str(i + 2))] = text
        except BaseException as be:
            pass
        i += 1
    wb.save(filename=filename)


if __name__ == "__main__":
    # with open('id3.txt', 'r') as file:
    #     lines = [line.replace('\n', '') for line in file.readlines()]
    filename = 'declaracii.xlsx'
    r = list(range(50, 81, 10))
    rm = 10
    with Pool() as pool:
        decla = pool.map(parse, [r])
        pool.close()
        pool.join()
