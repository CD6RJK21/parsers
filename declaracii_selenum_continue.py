from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import openpyxl
import time
from random import choice
from multiprocessing import Pool

import requests
from bs4 import BeautifulSoup


PROXY = ["198.50.251.188:808", "186.206.144.130:80", "87.255.70.183:8080", "217.113.122.142:3128"]

with open('id3.txt', 'r') as file:
    lines = [line.replace('\n', '') for line in file.readlines()]
filename = 'declaracii.xlsx'
wb = openpyxl.load_workbook(filename)

ws = wb['decla']

chrome_options = webdriver.ChromeOptions()
# chrome_options.add_argument('--proxy-server=http://%s' % choice(PROXY))
# chrome_options.add_argument("user-data-dir=selenium")
driver = webdriver.Chrome(chrome_options=chrome_options)

i = int(input())
m = int(input())
print(i, m)


while i <= m:
    chrome_options.add_argument('--proxy-server=http://%s' % choice(PROXY))
    line = lines[i]
    url = "https://pub.fsa.gov.ru/rds/declaration/view/{}/applicant".format(
        line)
    driver.get(url)
    try:
        text = WebDriverWait(driver, 2).until(
            EC.visibility_of_element_located(
                (By.XPATH,
                 "/html/body/fgis-root/div/fgis-rds-view-declaration/fgis-rds-view-declaration-toolbar/div"))).text
    except TimeoutException as te:
        time.sleep(5)
        continue
    try:
        text = driver.find_element_by_xpath('/html/body/fgis-root/div/fgis-rds-view-declaration/div/div/div/div/fgis-rds-view-application-applicant/fgis-card-block-wrapper/fgis-card-block/div/div[2]/div/fgis-card-info-row/div[2]/span').text
        ws['A{}'.format(str(i + 2))] = text
    except BaseException as be:
        pass
    try:
        text = driver.find_element_by_xpath(
            '/html/body/fgis-root/div/fgis-rds-view-declaration/div/div/div/div/fgis-rds-view-application-applicant/fgis-card-block-wrapper/fgis-card-block/div/div[2]/div/fgis-rds-view-contacts/fgis-card-row-spoiler/div[2]/fgis-card-edit-row-two-columns/fgis-card-info-row[1]/div[2]/p').text
        ws['B{}'.format(str(i + 2))] = text
    except BaseException as be:
        pass
    i += 1
wb.save(filename=filename)
